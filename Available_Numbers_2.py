import os
import arcpy
import datetime
import xlrd
import openpyxl as xl
from openpyxl import load_workbook
import pyexcel as p
import shutil

##### This creates a list of licences numbers from L1 to L60000 as control #####

AvailableNumbers = []
iterator = 1
for number in range(0, 60000):
    AvailableNumbers.append("L" + str(iterator))
    iterator += 1

AvailableORN = []
iterator = 1
for number in range(0, 60000):
    AvailableORN.append(iterator)
    iterator += 1

AvailableUNumbers = []
iterator = 1
for number in range(0, 60000):
    AvailableUNumbers.append("U" + str(iterator))
    iterator += 1


desktopPath = os.environ.get( "USERPROFILE" ) + "\\Desktop"

os.makedirs(desktopPath + "\\temp")

TempFolder = desktopPath + "\\temp"

in_feature_class = r"C:\Users\UKPXR011\Desktop\Scratch\Scratch.gdb\AccessLicence"

in_feature_class2 = r"C:\Users\UKPXR011\Desktop\Scratch\Scratch.gdb\Landownership"

##### This makes a table view of the licences features #####

arcpy.AddMessage("Creating Licence Table View")

arcpy.MakeTableView_management(in_feature_class, "AccessLicencesTV")

arcpy.AddMessage("Licence Table View Created")

out_xls_Licences = TempFolder + "\\LicencesTemp.xls"

##### This makes a table view of the ownership features #####

arcpy.AddMessage("Creating Ownership Table View")

arcpy.MakeTableView_management(in_feature_class2, "OwnershipTV")

arcpy.AddMessage("Ownership Table View Created")

out_xls_Ownership = TempFolder + "\\OwnershipTemp.xls"

##### This converts from table view to excel so the values can be extracted #####

arcpy.AddMessage("Converting Licence Table to Excel")

arcpy.TableToExcel_conversion("AccessLicencesTV", out_xls_Licences)

arcpy.AddMessage("Licence Table to Excel Converted")

##### This converts from table view to excel so the values can be extracted #####

arcpy.AddMessage("Converting Ownership Table to Excel")

arcpy.TableToExcel_conversion("OwnershipTV", out_xls_Ownership)

arcpy.AddMessage("Ownership Table to Excel Converted")

##### This coverts the .xls to .xlsx so it can be used with openpyxl (Licences) #####

p.save_book_as(file_name=out_xls_Licences,
               dest_file_name= TempFolder + "\\LicencesTemp_XLSX.xlsx")

arcpy.AddMessage("Excel Converted and Removing Duplicate Licence Values")

out_xlsx_Licences = TempFolder + "\\LicencesTemp_XLSX.xlsx"

##### This coverts the .xls to .xlsx so it can be used with openpyxl (Ownership) #####

p.save_book_as(file_name=out_xls_Ownership,
               dest_file_name= TempFolder + "\\OwnershipTemp_XLSX.xlsx")

arcpy.AddMessage("Excel Converted and Removing Duplicate Licence Values")

out_xlsx_Ownership = TempFolder + "\\OwnershipTemp_XLSX.xlsx"

##### This opens to .xlsx and extracts the values into a python list (Licences) #####

arcpy.AddMessage("Extracting Licence Values from .xlsx into python list")

Licencewb =   load_workbook(filename = out_xlsx_Licences)
Licencesh =   Licencewb['LicencesTemp']

Lnumbers = []
LicenceRows = Licencesh.max_row
iterator2 = 1
for row in range(0, LicenceRows):
    Lnumbers.append(str(Licencesh['B'+ str(iterator2)].internal_value))
    iterator2 += 1

##### This opens to .xlsx and extracts the values into a python list (Ownership) #####

arcpy.AddMessage("Extracting Ownership Values from .xlsx into python list")

Ownershipwb =   load_workbook(filename = out_xlsx_Ownership)
Ownershipsh =   Ownershipwb['OwnershipTemp']

ORNnumbers = []
OwnershipRows = Ownershipsh.max_row
iterator3 = 1
for row in range(0, OwnershipRows):
    ORNnumbers.append(Ownershipsh['S'+ str(iterator3)].internal_value)
    iterator3 += 1
UNumbers = []
for row in range(0, OwnershipRows):
    UNumbers.append(Ownershipsh['I'+ str(iterator3)].internal_value)
    iterator3 += 1

##### This compares the lists and removes duplicate values leaving just unique values in the list AvailableNumberFinal #####

arcpy.AddMessage("Comparing the Python Lists and Removing Duplicates")

AvailableNumberFinal = (list(set(AvailableNumbers) - set(Lnumbers))  )

AvailableORNNumbers = (list(set(AvailableORN) - set(ORNnumbers))  )

AvailableUNumbersFinal =  (list(set(AvailableUNumbers) - set(UNumbers))  )

AvailableNumberFinal.sort()

AvailableORNNumbers.sort()

AvailableUNumbersFinal.sort()

RangeEnd = len(AvailableNumberFinal)

RangeEnd2 = len(AvailableORNNumbers)

RangeEnd3 = len(AvailableUNumbersFinal)

wb = xl.Workbook()
ws = wb.active
for i in range(1, RangeEnd):
    ws['A' + str(i)] = AvailableNumberFinal[i]

for i in range(1, RangeEnd2):
    ws['B' + str(i)] = AvailableORNNumbers[i]

for i in range(1, RangeEnd3):
    ws['C' + str(i)] = AvailableUNumbersFinal[i]

arcpy.AddMessage("Duplicates Removed and AvailableNumbers.xlsx Produced in")


wb.save(desktopPath + "\\AvailableNumbers.xlsx")

shutil.rmtree(TempFolder)







